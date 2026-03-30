"""Endpoints CRUD para gestión de tablas en Tables.xlsx."""
import threading
import pandas as pd
from openpyxl import load_workbook, Workbook
from fastapi import APIRouter, Depends, HTTPException
from models.schemas import TableEntry, TableAdd, TableValidateRequest, TableValidateResult, TableBulkSave
from core.settings import settings
from core.alation_client import get_all
from dependencies import get_tokens

router = APIRouter()
_file_lock = threading.Lock()


def _read_tables() -> list[dict]:
    if not settings.tables_file.exists():
        return []
    df = pd.read_excel(settings.tables_file, sheet_name="Tables", engine="openpyxl")
    rows = []
    for _, row in df.iterrows():
        oid = int(row["Oid"])
        name = str(row["Name"]) if "Name" in df.columns and not pd.isna(row.get("Name")) else None
        rows.append({"oid": oid, "name": name})
    return rows


def _write_tables(entries: list[dict]) -> None:
    """Escribe la lista de OIDs en la hoja 'Tables' de Tables.xlsx."""
    tables_path = settings.tables_file
    tables_path.parent.mkdir(parents=True, exist_ok=True)

    if tables_path.exists():
        wb = load_workbook(tables_path)
        if "Tables" in wb.sheetnames:
            del wb["Tables"]
        ws = wb.create_sheet("Tables", 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tables"

    # Cabecera
    ws.cell(row=1, column=1, value="Oid")
    ws.cell(row=1, column=2, value="Name")

    # Datos
    for row_num, entry in enumerate(entries, start=2):
        ws.cell(row=row_num, column=1, value=entry["oid"])
        ws.cell(row=row_num, column=2, value=entry.get("name"))

    wb.save(tables_path)


@router.get("/", response_model=list[TableEntry])
async def list_tables():
    """Lista todas las tablas registradas en Tables.xlsx."""
    with _file_lock:
        return _read_tables()


@router.post("/", response_model=TableEntry, status_code=201)
async def add_table(body: TableAdd):
    """Agrega un OID de tabla a Tables.xlsx."""
    with _file_lock:
        entries = _read_tables()
        if any(e["oid"] == body.oid for e in entries):
            raise HTTPException(status_code=409, detail=f"El OID {body.oid} ya existe.")
        entries.append({"oid": body.oid, "name": None})
        _write_tables(entries)
    return TableEntry(oid=body.oid)


@router.post("/validate-ids", response_model=list[TableValidateResult])
async def validate_table_ids(
    body: TableValidateRequest,
    tokens: dict = Depends(get_tokens),
):
    """
    Consulta Alation para verificar si los OIDs existen y retorna sus nombres.
    Requiere Bearer Token válido del usuario.
    """
    base = settings.alation_base_url
    results: list[TableValidateResult] = []

    for oid in body.oids:
        try:
            data = get_all(
                f"{base}/catalog/table/",
                {"id": oid, "limit": 2},
                bearer_token=tokens["bearer_token"],
                api_token=tokens["api_token"],
            )
            if data:
                row = data[0]
                schema = row.get("schema_name") or ""
                name = row.get("name") or ""
                title = row.get("title") or ""
                full = f"{schema}.{name}" if schema else name
                results.append(TableValidateResult(
                    oid=oid,
                    exists=True,
                    name=name,
                    schema_name=schema,
                    full_name=full,
                    title=title,
                ))
            else:
                results.append(TableValidateResult(oid=oid, exists=False))
        except Exception:
            results.append(TableValidateResult(oid=oid, exists=False))

    return results


@router.post("/save-bulk", response_model=list[TableEntry], status_code=201)
async def save_bulk_tables(body: TableBulkSave):
    """
    Reemplaza completamente Tables.xlsx con la lista de entradas proporcionada.
    Útil para guardar OIDs validados desde el frontend.
    """
    if not body.entries:
        raise HTTPException(status_code=422, detail="La lista de entradas no puede estar vacía.")
    with _file_lock:
        entries = [{"oid": e.oid, "name": e.name} for e in body.entries]
        _write_tables(entries)
    return body.entries


@router.delete("/{oid}", status_code=204)
async def delete_table(oid: int):
    """Elimina un OID de tabla de Tables.xlsx."""
    with _file_lock:
        entries = _read_tables()
        new_entries = [e for e in entries if e["oid"] != oid]
        if len(new_entries) == len(entries):
            raise HTTPException(status_code=404, detail=f"OID {oid} no encontrado.")
        _write_tables(new_entries)
